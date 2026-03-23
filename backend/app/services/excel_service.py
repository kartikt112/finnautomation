import csv
from openpyxl import load_workbook


def parse_excel(file_path: str) -> list[dict]:
    """Parse an Excel or CSV file and return a list of row dicts.
    Validates that required columns (email/username + password) exist.
    """
    if file_path.endswith(".csv"):
        return _parse_csv(file_path)
    return _parse_xlsx(file_path)


def _parse_csv(file_path: str) -> list[dict]:
    """Parse a CSV file."""
    with open(file_path, "r", encoding="utf-8") as f:
        reader = csv.reader(f)
        rows = list(reader)

    if len(rows) < 2:
        raise ValueError("CSV file must have a header row and at least one data row")

    headers = [h.strip().lower() for h in rows[0]]
    return _process_rows(headers, rows[1:])


def _parse_xlsx(file_path: str) -> list[dict]:
    """Parse an Excel file."""
    wb = load_workbook(file_path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 2:
        raise ValueError("Excel file must have a header row and at least one data row")

    headers = [str(h).strip().lower() if h else "" for h in rows[0]]
    str_rows = [[str(cell) if cell is not None else "" for cell in row] for row in rows[1:]]
    return _process_rows(headers, str_rows)


def _process_rows(headers: list[str], rows: list) -> list[dict]:
    """Process rows with validated headers."""
    has_email = any(h in ("email", "username", "user", "login") for h in headers)
    has_password = "password" in headers

    if not has_email:
        raise ValueError("File must have an 'Email' or 'Username' column")
    if not has_password:
        raise ValueError("File must have a 'Password' column")

    # Build clean header names
    clean_headers = []
    for h in headers:
        if h in ("email", "username", "user", "login"):
            clean_headers.append("username")
        elif h == "2fa token":
            clean_headers.append("2fa_token")
        elif h == "link 2fa":
            clean_headers.append("link_2fa")
        elif h == "view inbox online":
            clean_headers.append("view_inbox")
        else:
            clean_headers.append(h.replace(" ", "_"))

    data = []
    for row in rows:
        if all(not cell or cell == "None" for cell in row):
            continue
        entry = {}
        for i, cell in enumerate(row):
            if i < len(clean_headers):
                val = str(cell).strip() if cell else ""
                entry[clean_headers[i]] = val
        if entry.get("username") and entry.get("password"):
            data.append(entry)

    return data
